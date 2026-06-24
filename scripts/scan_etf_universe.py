"""
REQ-274: ETF pool auto-screening pipeline.
Usage: python scripts/scan_etf_universe.py [--debug] [--force-refresh]
  --debug: generate outputs/etf_screening_debug.xlsx
"""
import sys, json, re, time, os, argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

PROJECT_ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "config").is_dir() and (parent / "scripts").is_dir())
DATA_DIR = PROJECT_ROOT / "data" / "quant"
SCREEN_DIR = PROJECT_ROOT / "data" / "screening"
VOL_CACHE = SCREEN_DIR / "step2_amount.json"
SINA_URL = "https://hq.sinajs.cn/list="
SINA_HDR = {"User-Agent": "Mozilla/5.0", "Referer": "https://finance.sina.com.cn/"}

parser = argparse.ArgumentParser()
parser.add_argument("--debug", action="store_true")
parser.add_argument("--force-refresh", action="store_true")
parser.add_argument("--dry-run", action="store_true", help="Skip Excel, print summary only")
args = parser.parse_args()

# ============================================================
# 1. Load ETF universe
# ============================================================
print("[1/4] Loading ETF universe...")
import akshare as ak, pandas as pd, numpy as np, yaml, requests

# Load/cache listing dates (Sina K-line first bar date)
_LISTING_CACHE = SCREEN_DIR / 'step3_listing_dates.json'
_listing_dates = {}
if _LISTING_CACHE.exists():
    _listing_dates = json.load(open(_LISTING_CACHE, 'r', encoding='utf-8'))

df = ak.fund_name_em()
c_col, n_col, t_col = df.columns[0], df.columns[2], df.columns[3]
ETF_RE = re.compile(r'^(159\d{3}|5[1-9]\d{4})$')
df = df[df[c_col].apply(lambda c: bool(ETF_RE.match(str(c))))].copy()
df.columns = ['code','pinyin','name','type','pf']
print(f"  {len(df)} ETF-likes")

# Load AUM cache (fund size from 天天基金, cached 168h)
# Replaces throttled fund_etf_spot_em() (Eastmoney push2).
# Actual fetch deferred to after Step 2 — only for codes with trading volume.
_SPOT_CACHE = SCREEN_DIR / 'step3_aum.json'
s_code = '代码'
s_cap  = '流通市值'
_AUM_MAX_AGE = 168  # 1 week (fund size updates quarterly, not daily)
_spot_age_h = (time.time() - os.path.getmtime(_SPOT_CACHE)) / 3600 if _SPOT_CACHE.exists() else 999
spot = pd.DataFrame()
if _SPOT_CACHE.exists():
    try:
        spot = pd.read_json(_SPOT_CACHE)
        print(f"  Loaded {len(spot)} cached AUM records ({_spot_age_h:.0f}h old)")
    except Exception as e:
        print(f"  [WARN] Failed to load AUM cache: {e}")
        spot = pd.DataFrame()

# ============================================================
# 2. Fetch volume + exchange names from Sina
# ============================================================
print("[2/4] Volume data...")
now = datetime.now()
# Only allow fetch after 15:10 (confirmed close data).
# During session (9:30-15:10) and weekends: ALWAYS use cache — mid-session
# accumulated volume is unreliable per REQ-274 screening rules.
is_post_close = now.hour > 15 or (now.hour == 15 and now.minute >= 10)
cache = {}; fetch_needed = True
if VOL_CACHE.exists() and not args.force_refresh:
    with open(VOL_CACHE) as f: cache = json.load(f)
    age_h = (time.time() - os.path.getmtime(VOL_CACHE)) / 3600
    codes_all = set(df['code'])
    missing_n = sum(1 for c in codes_all if c not in cache)
    if not is_post_close:
        # Pre-close: never fetch — cache holds last close data, still valid
        fetch_needed = False
        print(f"  Pre-close: using cache ({len(cache)} ETFs, {age_h:.1f}h old, last close data)")
    elif age_h < 24 and missing_n < len(codes_all) * 0.3:
        fetch_needed = False
        print(f"  Post-close cache valid ({len(cache)} ETFs, {age_h:.1f}h old)")

if fetch_needed:
    if not is_post_close and args.force_refresh:
        print(f"  *** REFUSING --force-refresh pre-close: mid-session data is unreliable ***")
        print(f"  *** Using cache instead. Re-run after 15:10 if fresh close data is needed. ***")
        with open(VOL_CACHE) as f: cache = json.load(f)
        fetch_needed = False
    else:
        print(f"  Fetching {len(df)} ETFs via Sina (post-close refresh)...")
    cache = {}
    cache['_meta'] = {'fetched_at': now.isoformat(), 'is_close_data': True}
    codes_sorted = sorted(df['code'])
    market = {c: ('sh' if c.startswith(('51','56','58','52','50')) else 'sz') for c in codes_sorted}
    symbols = [f"{market[c]}{c}" for c in codes_sorted]
    n_total = (len(symbols) + 49) // 50
    for i in range(0, len(symbols), 50):
        batch = symbols[i:i+50]; bn = i // 50 + 1
        try:
            url = f"{SINA_URL}{','.join(batch)}"
            resp = requests.get(url, headers=SINA_HDR, timeout=10)
            resp.encoding = 'gbk'
            for line in resp.text.strip().split('\n'):
                if '=' not in line: continue
                sm = re.match(r'var hq_str_(sh|sz)(\d{6})', line)
                if not sm: continue
                code = sm.group(2)
                qm = re.search(r'"([^"]*)"', line)
                if not qm: continue
                parts = qm.group(1).split(',')
                if len(parts) < 10: continue
                cache[code] = {
                    'name_ex': parts[0],
                    'open': float(parts[1] or 0), 'prev_close': float(parts[2] or 0),
                    'price': float(parts[3] or 0), 'vol': int(float(parts[8] or 0)),
                }
        except Exception as e:
            print(f"  Batch {bn}/{n_total} error: {e}")
        if bn < n_total: time.sleep(1.5)
    SCREEN_DIR.mkdir(exist_ok=True)
    with open(VOL_CACHE, 'w') as f: json.dump(cache, f, ensure_ascii=False)
    print(f"  Cached {len(cache)} ETFs")

# Attach data
df['name_ex'] = df['code'].apply(lambda c: cache.get(c, {}).get('name_ex', ''))
df['vol'] = df['code'].apply(lambda c: cache.get(c, {}).get('vol', 0))
df['price'] = df['code'].apply(lambda c: cache.get(c, {}).get('price', 0))
df['amount'] = df['vol'] * df['price']
df['amount_yi'] = (df['amount'] / 1e8).round(2)

# Data length: how many trading days of history exist for each ETF.
# Truth source: Sina historical K-line API (independent of quant/backtest flow).
# Local 回测日线 CSV (data/quant/{code}_daily.csv) is only used as a cache —
# when present we can skip the API call, but the data is the same external source.
# Returns -1 as sentinel for "not checked yet" (resolved below via Sina).
MIN_HISTORY_DAYS = 80  # F7 factor requires 80 daily bars minimum

def _history_days(code):
    """Return known trading-day count for an ETF code.
    Checks local 回测日线 cache first (fast), returns -1 if not cached."""
    kline_csv = DATA_DIR / f'{code}_daily.csv'  # 回测日线 (quant flow output, same Sina data)
    if kline_csv.exists():
        try:
            return len(pd.read_csv(kline_csv))
        except:
            pass
    return -1  # sentinel: cache miss, needs Sina API fetch below

df['history_days'] = df['code'].apply(_history_days)
n_csv = (df['history_days'] >= 0).sum()
n_missing = (df['history_days'] == -1).sum()
print(f"  Data check: {n_csv} from CSV, {n_missing} need API check")

# For ETFs not in local cache, check history_days cache first, then fallback to Sina API.
_HISTORY_CACHE = SCREEN_DIR / 'history_days.json'
_history_map = {}
if _HISTORY_CACHE.exists():
    try:
        _history_map = json.load(open(_HISTORY_CACHE, 'r', encoding='utf-8'))
        # Use cached values for known codes
        for _code, _rows in _history_map.items():
            if _code in df['code'].values and df.loc[df['code'] == _code, 'history_days'].values[0] == -1:
                df.loc[df['code'] == _code, 'history_days'] = _rows
        print(f"  Loaded {len(_history_map)} history_days from cache")
    except Exception:
        _history_map = {}

# Remaining unfetched ETFs that pass amount filter → fetch from Sina
_rows_to_fetch = [c for c in df['code']
                  if df.loc[df['code'] == c, 'history_days'].values[0] == -1
                  and c in cache
                  and cache[c].get('vol', 0) * cache[c].get('price', 0) >= 30_000_000]
if _rows_to_fetch:
    print(f"  Fetching data row counts for {len(_rows_to_fetch)} ETFs via Sina...")
    SINA_HIST_URL = 'https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/CN_MarketData.getKLineData'
    SINA_HDR = {"User-Agent": "Mozilla/5.0", "Referer": "https://finance.sina.com.cn/"}
    fetched_rows = 0

    def _fetch_rows(code):
        market = 'sh' if code.startswith(('51','56','58','52','50')) else 'sz'
        try:
            r = requests.get(SINA_HIST_URL,
                           params={'symbol': f'{market}{code}', 'scale': 240, 'ma': 'no', 'datalen': 5000},
                           headers=SINA_HDR, timeout=8)
            r.encoding = 'gbk'
            data = json.loads(r.text)
            return code, len(data) if isinstance(data, list) else 0
        except:
            return code, 0

    with ThreadPoolExecutor(max_workers=3) as pool:  # conservative to avoid IP ban
        futures = {pool.submit(_fetch_rows, c): c for c in _rows_to_fetch}
        for f in as_completed(futures):
            code, rows = f.result()
            df.loc[df['code'] == code, 'history_days'] = rows
            _history_map[code] = rows
            fetched_rows += 1
            if fetched_rows % 100 == 0:
                print(f'    {fetched_rows}/{len(_rows_to_fetch)}')
    # Save cache (new codes appended)
    json.dump(_history_map, open(_HISTORY_CACHE, 'w', encoding='utf-8'), ensure_ascii=False)
    print(f'    Done: {fetched_rows} fetched, cache saved')
# Mark remaining unfetched as 0 (will be filtered by minimum)
df.loc[df['history_days'] == -1, 'history_days'] = 0

n_below = (df['history_days'] < MIN_HISTORY_DAYS).sum()
print(f"  Final: {(df['history_days'] >= MIN_HISTORY_DAYS).sum()} ETFs >= {MIN_HISTORY_DAYS}d, {n_below} below")

# Load AUM data (fund size from 天天基金 cache)
_spot_map = {}
if len(spot) > 0:
    for _, r in spot.iterrows():
        try:
            sc = str(int(r[s_code])).zfill(6)
            _spot_map[sc] = float(r[s_cap]) / 1e8 if pd.notna(r[s_cap]) else 0
        except: pass
# AUM: use cached data. Missing codes that pass amount filter will be fetched below.
df['mcap_yi'] = df['code'].apply(lambda c: _spot_map.get(c, 0))

# Proactively fill AUM gaps: fetch missing codes via 天天基金 API.
_aum_missing = [c for c in df['code']
                if c not in _spot_map
                and c in cache
                and cache[c].get('vol', 0) * cache[c].get('price', 0) >= 30_000_000]
if _aum_missing:
    print(f'  Fetching AUM for {len(_aum_missing)} ETFs...')
    try:
        sys.path.insert(0, str(PROJECT_ROOT))
        from _working.fetch_aum_ttjj import fetch_batch as _aum_fetch
        _new = _aum_fetch(_aum_missing, delay=0.35)
        for k, v in _new.items():
            _spot_map[k] = v
            df.loc[df['code'] == k, 'mcap_yi'] = v
        print(f'    Fetched {len(_new)} AUM records')
    except Exception as e:
        print(f'    [WARN] AUM fetch failed: {e} — {len(_aum_missing)} ETFs have AUM=0')


# ============================================================
# 3. Combined filter
# ============================================================
print("[3/4] Filtering...")
BAD_TYPES = ('固收','货币','债')  # type field contains these -> exclude

def passes_filter(r):
    if r['amount'] < 30_000_000: return False, 'amount<30M'
    if r['history_days'] < MIN_HISTORY_DAYS: return False, f'history<{MIN_HISTORY_DAYS}d'
    ftype = str(r['type'])
    for kw in BAD_TYPES:
        if kw in ftype: return False, 'excluded: ' + kw
    return True, ''

reasons = []
keep_mask = []
for _, r in df.iterrows():
    ok, reason = passes_filter(r)
    reasons.append(reason)
    keep_mask.append(ok)

df['filter_reason'] = reasons
df['_keep'] = keep_mask
df_filtered = df[df['_keep']].copy()
df_filtered = df_filtered.drop(columns=['_keep'])
df = df.drop(columns=['_keep'])
print(f"  {len(df)} -> {len(df_filtered)} (removed {len(df)-len(df_filtered)})")

# ============================================================
# 4. Classification + Holdings dedup + Selection
# ============================================================
print("[4/4] Classifying & selecting...")

# Load holdings (independent cache, NOT pool data)
_HCACHE = SCREEN_DIR / 'step4_holdings.json'
holdings_data = json.load(open(_HCACHE, 'r', encoding='utf-8')) if _HCACHE.exists() else {}
def hs(code):
    top10 = holdings_data.get(code, [])
    return {h['code'] for h in top10 if 'code' in h}
def exch(code):
    return 'sh' if code.startswith(('51','56','58','52','50')) else 'sz'

# Name groups — must cover ALL ETFs, zero "其他"
# Ordering is critical: more specific patterns first, catchalls later.
# HK groups before A-share; 消费电子 before 电子; 金融科技 before 科技.
NAME_GROUPS = [
    # === QDII by country/region ===
    # US: sector keywords first, then broad-market catchall (treated like 宽基)
    ('QD-美国-科技', ['纳指科技','纳斯达克科技']),
    ('QD-美国-油气', ['标普.*油气','标普.*原油']),
    ('QD-美国-生物', ['标普.*生物','标普.*生科']),
    ('QD-美国-消费', ['标普.*消费']),
    ('QD-美国-宽基', ['纳指','标普','道琼斯','美国.*50','纳斯达克']),
    ('QD-日本', ['日经','东证']),
    ('QD-韩国', ['韩国','中韩']),
    ('QD-德国', ['德国']), ('QD-法国', ['法国']), ('QD-沙特', ['沙特']),
    ('QD-巴西', ['巴西']), ('QD-印度', ['印度']), ('QD-越南', ['越南']),
    ('QD-东南亚', ['东南亚']), ('QD-亚太', ['亚太']),
    ('HK-中概互联', ['中概.*互联','海外.*互联','中国互联']),
    # === HK by theme ===
    ('HK-医药', ['港股.*药','香港.*药','恒生.*药','恒生.*医','港股通.*药','港股通.*医']),
    ('HK-金融', ['港股.*券','香港.*券','港股.*非银','香港.*非银','港股.*金融','香港.*金融']),
    ('HK-银行', ['港股.*银行','香港.*银行','恒生.*银行']),
    ('HK-红利', ['港股红利','香港红利','恒生红利','港股通红利','港股通.*红利','恒生.*红利','港股.*低波','香港.*低波','恒生.*低波']),
    ('HK-信息技术', ['港股.*信息','港股通.*信息','香港.*信息','恒生.*信息']),
    ('HK-互联网', ['港股.*互联网','港股通.*互联网','香港.*互联网','恒生.*互联网','中概.*互联','中国互联']),
    ('HK-消费', ['港股.*消费','香港.*消费','恒生.*消费','港股通.*消费']),
    ('HK-科技', ['港股.*科','香港.*科','恒生.*科','港股通.*科','港股通.*新经济']),
    ('HK-房地产', ['港股.*地产','香港.*地产','恒生.*地产']),
    ('HK-汽车', ['港股.*汽车','香港.*汽车']),
    ('HK-能源', ['港股.*能源','香港.*能源']),
    ('HK-高股息', ['港股.*高股息','港股通.*高股息','恒生.*高股息','港股.*高息','港股通.*高息']),
	    ('HK-国企', ['港股.*国企','恒生.*国企','港股通.*国企']),
    ('HK-宽基', ['恒生ETF','恒生指数','港股通ETF','港股通指数','港股通.*50','港股通.*100','港股.*100','香港.*30','恒生']),
    # === A-share: industry sectors (specific sub-groups first) ===
    ('消费电子', ['消费电子','消电']),
    ('芯片', ['芯片','半导体','集成电路','电子']),
    ('医药-创新药', ['创新药','生物医药']),
    ('医药-中药', ['中药']),
    ('医药-器械', ['医疗器械','医械','医疗设备']),
    ('医药', ['医药','医疗','生物','疫苗']),
    ('金融-保险', ['保险']), ('金融-证券', ['证券','券商']),
    ('金融-银行', ['银行']), ('金融-非银', ['非银']),
    ('金融科技', ['金融科技']),
    ('金融', ['金融']),  # broad financial catchall (after all specific subgroups)
    ('消费-酒', ['酒']),
    ('消费-食品', ['食品','饮料','国货']),
    ('消费', ['消费']),
    ('消费-家电', ['家电']), ('消费-旅游', ['旅游']),
    ('新能源-光伏', ['光伏']), ('新能源-电池', ['电池']),
    ('新能源-电力', ['电力']), ('新能源-电网', ['电网']),
    ('新能源-风电', ['风电']), ('新能源-储能', ['储能']),
    ('新能源', ['新能源']),
    ('军工', ['军工','国防','航空','通用航空']), ('传媒', ['传媒']), ('游戏', ['游戏']),
    ('煤炭', ['煤炭']), ('钢铁', ['钢铁']), ('化工', ['化工']),
    ('工业母机', ['工业母机','机床']),
    ('机械', ['机械','工程机械']), ('石油', ['石油','石化','油气','原油']),
    ('房地产', ['地产','房地产']),
    ('农业-养殖', ['养殖','畜牧']),
    ('农业-种植', ['粮食','种植','种业']),
    ('农业', ['农业','农牧','农产品','现代农业']),
    ('有色-稀土', ['稀土','稀有金属','稀金']),
    ('有色', ['有色','矿业','工业金属']),
    ('黄金', ['黄金','金ETF','上海金']), ('豆粕', ['豆粕']),
    ('通信', ['通信','5G','电信']),
    ('计算机AI', ['计算机','软件','大数据','云计算','数字经济','信创','人工智能','AI','科技','信息','VR']),
    ('汽车', ['汽车','新能源车','智能车','智能驾驶']),
    ('机器人', ['机器人']), ('基建建材', ['基建','建材']),
    ('电力公用', ['电力','公用事业']),
    ('物流运输', ['运输','物流']),
    ('环保碳中和', ['环保','碳中和','绿电']),
    ('卫星航天', ['卫星','航天']), ('船舶', ['船舶']),
    ('动漫游戏', ['动漫']), ('影视', ['影视']),
    ('新材料', ['新材料']),
    ('资源', ['资源','大宗商品']), ('能源', ['能源']),
    ('高端装备', ['高端装备','智能制造','战略新兴']),
    # === Financial / Style / Thematic (catchall order) ===
    ('红利', ['红利']),
    ('高股息', ['高股息','高息']),  # A-share high-dividend (HK version above)
    ('现金流', ['自由现金流','现金流']),
    ('央企国企', ['央企','国企','国企改革']),
    ('一带一路', ['一带一路']), ('专精特新', ['专精特新']),
    ('ESG', ['ESG','可持续发展','责任投资']), ('沪深港', ['沪深港']),
    ('教育', ['教育']), ('区域经济', ['经济圈','成渝','都市经济圈']),
    # === Style (very broad, must be AFTER all sector groups) ===
    ('风格', ['成长','价值','质量']),
    # === Broad-market (excluded except micro-cap 563300) ===
    ('宽基指数', ['沪深300','上证50','中证500','中证1000','中证2000','创业板',
                '科创50','科创100','科创200','深证100','深100','上证180','A50',
                'MSCI','中证100','A100','A500','深证50','上证综指','中证全指',
                '国证2000','科创综指','科创创业','上证指数','中小100','中证800',
                '双创50','创业大盘','TMT50','TMT','央企50','红利低波100',
                '深成','深证成指','深证基本面','基本面','纳斯达克','创50','国证','A股','科创增强','创新100']),
]

def classify(row):
    name = str(row['name']); code = str(row['code'])
    for grp, kws in NAME_GROUPS:
        for kw in kws:
            if re.search(kw, name):
                if grp == '宽基指数' and code == '563300':
                    return '微盘'
                return grp
    # Fallback: use type field (should be rare after comprehensive NAME_GROUPS)
    ftype = str(row['type'])
    return '其他-' + ftype.replace('型-','-')

df_filtered['cg'] = df_filtered.apply(classify, axis=1)
# Load pool codes early (needed for broad-market exemption)
with open(PROJECT_ROOT / 'config/quant_universe.yaml', 'r', encoding='utf-8') as f:
    cfg = yaml.safe_load(f)
pool_codes = {e['code'] for e in cfg['universe']}
# Remove broad-market ETFs (except pool ETFs, 微盘/563300, and alpha-potential keywords)
# QD-美国-宽基 = US broad-market (纳指/标普/道琼斯) treated like A-share 宽基指数
BROAD_CGS = {'宽基指数', '风格', 'QD-美国-宽基'}
# Keywords for broad-market ETFs that show alpha characteristics — let through to compete
BROAD_ALPHA_KWS = ['科创50', '科创100', '创业板指']
def _has_alpha_kw(row):
    name = str(row['name'])
    return any(kw in name for kw in BROAD_ALPHA_KWS)
df_filtered['_is_broad'] = (df_filtered['cg'].isin(BROAD_CGS)
    & (~df_filtered['code'].isin(pool_codes))
    & (df_filtered['code'] != '563300')
    & (~df_filtered.apply(_has_alpha_kw, axis=1)))
n_broad = df_filtered['_is_broad'].sum()
df_filtered = df_filtered[~df_filtered['_is_broad']]

# Count unclassified
n_other = (df_filtered['cg'].str.startswith('其他-')).sum()
print(f"  Groups: {df_filtered['cg'].nunique()} (removed {n_broad} broad/style ETFs)")
if n_other == 0:
    print(f"  *** ZERO unclassified ETFs ***")
else:
    print(f"  *** WARNING: {n_other} unclassified ETFs remain ***")
    for _, r in df_filtered[df_filtered['cg'].str.startswith('其他-')].iterrows():
        print(f"    {r['code']} | {r['name']} | type={r['type']}")

# ============================================================
# Step 3: Top-1 per coarse group (by daily turnover)
# ============================================================
import math
def etf_score(amt_yi, rows, aum_yi=0.01):
    # R13 weighting: turnover 35%, data length 50%, fund size 15%
    return (0.35 * math.log10(max(amt_yi, 0.01))
          + 0.50 * math.log10(max(rows, 1))
          + 0.15 * math.log10(max(aum_yi, 0.01)))

_aum_map = {}
for _code in df_filtered['code']:
    _aum = df_filtered[df_filtered['code'] == _code]['mcap_yi'].values[0]
    _aum_map[_code] = _aum if _aum > 0 else 0.01

df_filtered['log_score'] = df_filtered.apply(
    lambda r: etf_score(r['amount_yi'], r['history_days'], _aum_map.get(r['code'], 0.01)), axis=1)
df_filtered = df_filtered.sort_values('log_score', ascending=False)
top1_per_cg = []
for cg, grp in df_filtered.groupby('cg'):
    top1_per_cg.append(grp.iloc[0].to_dict())
top1_per_cg.sort(key=lambda x: -x['log_score'])

# ============================================================
# Step 4: Greedy overlap reduction
# Iteratively remove the ETF that shares the most duplicate stocks,
# until no stock appears in >= 7 selected ETFs.
# QDII/HK/commodity groups are exempt from elimination.
EXEMPT_PREFIXES = ('QD-',)  # Only QDII exempt from overlap elimination (HK ETFs participate)
EXEMPT_CGS = {'黄金', '豆粕'}  # 商品 ETF，无 A 股成分股。油气已移除（含 A 股 ETF）

# Separate exempt (always kept) from candidates (may be eliminated)
exempt = [s for s in top1_per_cg if s['cg'].startswith(EXEMPT_PREFIXES) or s['cg'] in EXEMPT_CGS]
candidates = [s for s in top1_per_cg if s not in exempt]

# Pre-load all A-share holdings
_cand_holdings = {}
for s in candidates:
    _h = hs(s['code'])
    if _h: _cand_holdings[s['code']] = _h

from collections import Counter
def _overlap_score(code, holdings_count):
    """Count how many of this ETF's holdings are duplicates (appear >=2 times)."""
    if code not in _cand_holdings: return 0
    return sum(1 for h in _cand_holdings[code] if h in holdings_count and holdings_count[h] >= 2)

# Build initial counts
holdings_count = Counter()
for s in exempt + candidates:
    if s['code'] in _cand_holdings:
        holdings_count.update(_cand_holdings[s['code']])

removed_pass1 = []  # bottom 1/2, overlap > 5 (O1)
removed_pass2 = []  # greedy, max_overlap >= 7 (O2)

# First pass: stricter overlap check on bottom 1/2 by log_score
if candidates:
    candidates.sort(key=lambda s: s.get('log_score', 0))  # low to high
    cutoff = len(candidates) // 2
    low_quartile = candidates[:cutoff]
    for s in low_quartile:
        if s not in candidates: continue
        if _overlap_score(s['code'], holdings_count) > 5:
            candidates.remove(s)
            removed_pass1.append(s['code'])
            holdings_count = Counter()
            for _s in exempt + candidates:
                if _s['code'] in _cand_holdings:
                    holdings_count.update(_cand_holdings[_s['code']])

# Second pass: greedy reduction on remaining candidates until max_overlap < 7
while True:
    scores = [(s, _overlap_score(s['code'], holdings_count)) for s in candidates]
    max_score = max(s[1] for s in scores) if scores else 0
    if max_score < 7:
        break
    worst = max(scores, key=lambda x: (x[1], -x[0].get('log_score', 0)))
    worst_etf = worst[0]
    candidates.remove(worst_etf)
    removed_pass2.append(worst_etf['code'])
    holdings_count = Counter()
    for s in exempt + candidates:
        if s['code'] in _cand_holdings:
            holdings_count.update(_cand_holdings[s['code']])

kept = exempt + candidates

# Print overlap statistics
if 'holdings_count' in dir():
    _dist = Counter()
    for _h, _c in holdings_count.items():
        if _c >= 2: _dist[_c] += 1
    print(f"  Overlap after reduction: {sum(_dist.values())} duplicate stocks across {len(kept)} ETFs")
    for _k in sorted(_dist.keys(), reverse=True):
        print(f"    overlap={_k}: {_dist[_k]} stocks")

deduped = pd.DataFrame(kept) if kept else pd.DataFrame()
sel_codes = set(deduped['code']) if len(deduped) > 0 else set()

# Mark selection in the filtered dataframe
df_filtered['fg'] = df_filtered['cg']
df_filtered['selected'] = df_filtered['code'].apply(lambda c: 'YES' if c in sel_codes else '')
df_filtered['in_pool'] = df_filtered['code'].apply(lambda c: 'YES' if c in pool_codes else '')
overlap = sel_codes & pool_codes
missing = pool_codes - sel_codes

# Count by category
n_qdii = sum(1 for s in kept if s['cg'].startswith('QD-'))
n_hk = sum(1 for s in kept if s['cg'].startswith('HK-'))
n_cmdty = sum(1 for s in kept if s['cg'] in EXEMPT_CGS)
n_a = len(kept) - n_qdii - n_hk - n_cmdty

print(f"  Step 3: {len(top1_per_cg)} ETFs (top-1 per coarse group)")
print(f"  Step 4: {len(kept)} ETFs after overlap reduction (①{len(removed_pass1)} + ②{len(removed_pass2)}, O1=5 O2=7)")
print(f"  Breakdown: QDII={n_qdii}, HK={n_hk}, Commodity={n_cmdty}, A-share={n_a}")
_new_codes_console = sel_codes - pool_codes
print(f"  差异: 重叠{len(overlap)} | 新增{len(_new_codes_console)} | 未入选{len(missing)}")
if missing:
    print(f"  未入选: {sorted(missing)}")
if _new_codes_console:
    _preview = sorted(_new_codes_console)[:12]
    print(f"  新增: {_preview}{'...' if len(_new_codes_console) > 12 else ''}")

# 5. Debug xlsx
# ============================================================
if args.dry_run:
    print("\n[Dry-run] skipping xlsx.")
    sys.exit(0)

if args.debug:
    print("\n[DEBUG] Generating xlsx...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OUTPUT = PROJECT_ROOT / "outputs" / "etf_screening_debug.xlsx"
    OUTPUT.parent.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HDR_FONT = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    RULE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    RULE_FONT = Font(name="Consolas", size=9, color="92400E")
    PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    GRP_FILLS = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in
                 ['DBEAFE','EDE9FE','FCE7F3','E0E7FF','FEF3C7','D1FAE5','FEE2E2']]
    thin_border = Border(left=Side(style='thin', color='E5E7EB'),
                         right=Side(style='thin', color='E5E7EB'),
                         top=Side(style='thin', color='E5E7EB'),
                         bottom=Side(style='thin', color='E5E7EB'))

    def rule_row(ws, row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = RULE_FILL; c.font = RULE_FONT; c.alignment = Alignment(wrap_text=True)
        for i in range(1, ncols+1): ws.cell(row=row, column=i).fill = RULE_FILL

    def hdr_row(ws, row, cols):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=col)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(horizontal='center')
            c.border = thin_border

    def write_rows(ws, start_row, df_data, cols, group_col=None):
        for idx, (_, r) in enumerate(df_data.iterrows()):
            row = start_row + idx
            for ci, col in enumerate(cols, 1):
                val = r.get(col, '');
                if isinstance(val, float): val = round(val, 2)
                c = ws.cell(row=row, column=ci, value=val)
                c.font = Font(name="Consolas", size=9); c.border = thin_border
            if group_col and group_col in r:
                ghash = hash(str(r[group_col])) % len(GRP_FILLS)
                for ci in range(1, len(cols)+1):
                    ws.cell(row=row, column=ci).fill = GRP_FILLS[ghash]

    # Sheet 1: Full universe
    ws1 = wb.create_sheet("1-全市场ETF")
    cols1 = ['code','name','type','amount_yi']
    rule_row(ws1, 1, '全市场 ETF (fund_name_em + Sina 成交数据). 按成交额降序.', len(cols1))
    hdr_row(ws1, 2, ['代码','名称','类型','成交额(亿)'])
    df1 = df.sort_values('amount', ascending=False)
    write_rows(ws1, 3, df1, cols1, group_col='type')
    for c, w in zip('ABCD', [10,32,24,14]): ws1.column_dimensions[c].width = w


    # Add broad_excluded to original df for Sheet 2
    BROAD_CGS_SET = {'宽基指数', '风格', 'QD-美国-宽基'}
    df['broad_excluded'] = df.apply(
        lambda r: 'YES' if (classify(r) in BROAD_CGS_SET and r['code'] != '563300' and r['code'] not in pool_codes) else '', axis=1
    )

    # Sheet 2: Combined filter (keep all rows, mark excluded in red)
    ws2 = wb.create_sheet("2-基础过滤")
    cols2 = ['code','name','type','amount_yi','history_days','filter_reason','broad_excluded']
    rule_row(ws2, 1, '基础过滤: 成交额<1000万 + 固收/货币/债券型 → filter_reason. 宽基/风格 → broad_excluded.', len(cols2))
    rule_row(ws2, 2, f'{len(df)} ETFs, 基础过滤剔除{len(df)-len(df_filtered)-n_broad}支, 宽基/风格排除{n_broad}支, 保留{len(df_filtered)}支', len(cols2))
    hdr_row(ws2, 3, ['代码','名称','类型','成交额(亿)','历史交易日数','基础排除原因','宽基/风格排除'])
    for idx, (_, r) in enumerate(df.sort_values('amount', ascending=False).iterrows()):
        row = 4 + idx
        for ci, col in enumerate(cols2, 1):
            val = r.get(col, '');
            if isinstance(val, float): val = round(val, 2)
            c = ws2.cell(row=row, column=ci, value=val)
            c.font = Font(name="Consolas", size=9); c.border = thin_border
        if r['filter_reason']:
            for ci in range(1, len(cols2)+1):
                ws2.cell(row=row, column=ci).fill = FAIL_FILL
    for c, w in zip('ABCDEFG', [10,32,24,14,14,28,16]): ws2.column_dimensions[c].width = w

    # Sheet 3: ALL ETFs per coarse group, sorted by score. Top-1 marked with ★.
    ws3 = wb.create_sheet("3-粗组排名")
    cols3 = ['code','name','type','amount_yi','history_days','aum_yi','log_score','cg','top1']  # history_days = 历史交易日数 (Sina API)
    rule_row(ws3, 1, 'Step 3: 加权得分 = 0.35×log10(成交额亿) + 0.50×log10(数据行数) + 0.15×log10(规模亿)。负分=低活跃/新上市/规模小。★ = 组内第1名。', len(cols3))
    rule_row(ws3, 2, f'筛选：成交额≥3000万 + 排除债/货币/固收类。{df_filtered["cg"].nunique()}个粗组, {len(df_filtered)}支ETF, 每组第1名共{len(top1_per_cg)}支进入Step4', len(cols3))
    hdr_row(ws3, 3, ['代码','名称','类型','成交额(亿)','历史交易日数','规模(流通市值亿)','加权得分','粗组','Top1'])
    df_cg_sorted = df_filtered.sort_values(['cg','log_score'], ascending=[True,False])
    cg_top_codes = {s['code'] for s in top1_per_cg}
    row_idx = 4
    for _, r in df_cg_sorted.iterrows():
        is_top1 = '★' if r['code'] in cg_top_codes else ''
        rows_val = r.get('history_days', 0)
        aum_val = round(_aum_map.get(r['code'],0), 2)
        vals = [r['code'], r['name'], r['type'], r['amount_yi'], rows_val,
                aum_val, round(r.get('log_score',0),3), r['cg'], is_top1]
        for ci, val in enumerate(vals, 1):
            if isinstance(val, float): val = round(val, 2)
            c = ws3.cell(row=row_idx, column=ci, value=val)
            c.font = Font(name="Consolas", size=9); c.border = thin_border
        if is_top1:
            for ci in range(1, len(cols3)+1):
                ws3.cell(row=row_idx, column=ci).fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        row_idx += 1
    for c, w in zip('ABCDEFGHI', [10,32,24,14,12,10,10,22,8]): ws3.column_dimensions[c].width = w

    # Sheet 4: Top-1 ETFs with Chinese holdings + Jaccard PK pairs
    ws4 = wb.create_sheet("4-成分股PK")
    # Pre-fetch missing A-share holdings before xlsx generation
    _missing = [s['code'] for s in top1_per_cg
                if not holdings_data.get(s['code'])
                and not (s['cg'].startswith('QD-') or s['cg'] in EXEMPT_CGS)]
    if _missing:
        print(f"  Fetching holdings for {len(_missing)} new A-share ETFs...")
        # Use a temp meta dict just for the fetch (not pool's etf_metadata.json)
        from fetch_etf_metadata import fetch_one
        _tmp_meta = {}
        for _code in _missing:
            try:
                _etf = {'code': _code, 'name': '', 'market': exch(_code), 'sector': ''}
                _result = fetch_one(_etf, _tmp_meta)
                holdings_data[_code] = _result.get('top10', [])
            except Exception as e:
                print(f'    {_code} FAIL: {e}')
        json.dump(holdings_data, open(_HCACHE, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        print(f'    Saved holdings for {len(_missing)} new ETFs')

    def get_holding_names(etf_code):
        top10 = holdings_data.get(etf_code, [])
        return [h.get('name', h.get('code','')) for h in top10][:10]
    # Build PK data
    pk_pairs = []
    # Only exclude QDII and commodities — HK ETFs participate (their HK-stock
    # holdings can overlap with each other, even if they don't overlap with A-shares).
    PK_EXEMPT_CGS = {'黄金', '豆粕'}
    for i, s1 in enumerate(top1_per_cg):
        cg1 = s1['cg']
        if cg1.startswith('QD-') or cg1 in PK_EXEMPT_CGS: continue
        for j, s2 in enumerate(top1_per_cg):
            if j <= i: continue
            cg2 = s2['cg']
            if cg2.startswith('QD-') or cg2 in PK_EXEMPT_CGS: continue
            if exch(s1['code']) != exch(s2['code']): continue
            sa, sb = hs(s1['code']), hs(s2['code'])
            if not sa or not sb: continue
            jac = round(len(sa & sb) / max(1, len(sa | sb)), 3)
            if jac > 0.3:
                pk_pairs.append({'code1':s1['code'],'name1':s1['name'],'cg1':cg1,'amt1':s1['amount_yi'],
                    'code2':s2['code'],'name2':s2['name'],'cg2':cg2,'amt2':s2['amount_yi'],'jaccard':jac})

    # Sort: A-share first, then HK, then others (QDII/commodity)
    def _cat_order(cg):
        if cg.startswith('QD-') or cg in EXEMPT_CGS: return 2
        if cg.startswith('HK-'): return 1
        return 0
    top1_sorted = sorted(top1_per_cg, key=lambda s: (_cat_order(s['cg']), s['cg'], -s['amount_yi']))

    # Build eliminated set (for coloring exclusion)
    _kept_codes = {s['code'] for s in kept}
    _eliminated = {s['code']: 'YES' for s in top1_per_cg if s['code'] not in _kept_codes}

    # Find duplicate holdings across A-share + HK ETFs (exclude eliminated ones)
    ASHARE_HK = [s for s in top1_sorted if _cat_order(s['cg']) < 2 and s['code'] in _kept_codes]
    from collections import Counter
    _holdings_count = Counter()
    for s in ASHARE_HK:
        _is_hk = s['cg'].startswith('HK-')
        for h in get_holding_names(s['code']):
            if h:
                key = h + (' -H' if _is_hk else '')
                _holdings_count[key] += 1
    _dup_names = {h for h, c in _holdings_count.items() if c >= 2}
    _dup_colors = ['FFB3BA','FFDFBA','FFFFBA','BAFFC9','BAE1FF','E8BAFF','FFB3E0','B3FFE0',
                   'FFD700','FF8C69','87CEEB','98FB98','DDA0DD','F0E68C','87CEFA','FFA07A']
    _dup_sorted = sorted(_dup_names, key=lambda n: -_holdings_count[n])
    _dup_fills = {n: PatternFill(start_color=_dup_colors[i % 16], end_color=_dup_colors[i % 16], fill_type='solid')
                  for i, n in enumerate(_dup_sorted)}

    # Build removal order maps from both passes (emoji-prefixed)
    _removed_order = {}
    for _i, _code in enumerate(removed_pass1, 1):
        _removed_order[_code] = f'①{_i}'  # ①1, ①2...
    for _i, _code in enumerate(removed_pass2, 1):
        _removed_order[_code] = f'②{_i}'  # ②1, ②2...

    _total_removed = len(removed_pass1) + len(removed_pass2)

    # Build separate A-share and HK internal rankings
    _a_ranked = sorted(
        [s for s in top1_per_cg if _cat_order(s['cg']) == 0],
        key=lambda s: -s.get('log_score', 0)
    )
    _h_ranked = sorted(
        [s for s in top1_per_cg if _cat_order(s['cg']) == 1],
        key=lambda s: -s.get('log_score', 0)
    )
    _a_rank = {s['code']: f'A{i+1}' for i, s in enumerate(_a_ranked)}
    _h_rank = {s['code']: f'H{i+1}' for i, s in enumerate(_h_ranked)}
    _ashare_cutoff = len(_a_ranked) // 2  # bottom 1/2 threshold for overlap pass 1

    # Combined rank map for elimination tracking (use sequential index for pass logic)
    _all_ranked = _a_ranked + _h_ranked
    _all_rank = {s['code']: i+1 for i, s in enumerate(_all_ranked)}

    # Main table
    HCOLS = ['持仓1','持仓2','持仓3','持仓4','持仓5','持仓6','持仓7','持仓8','持仓9','持仓10']
    cols4_header = ['代码','名称','粗组','成交额(亿)','得分','排名','排除'] + HCOLS
    _desc = f'Step 4: ①底1/2(A股{_ashare_cutoff}支) O1>5→淘汰({len(removed_pass1)}支) ②贪心 O2>=8→淘汰({len(removed_pass2)}支). 结果: {len(_kept_codes)}/{len(top1_per_cg)} 入选.'
    rule_row(ws4, 1, _desc, len(cols4_header))
    rule_row(ws4, 2, '排名: A1~A' + str(len(_a_ranked)) + '=A股 | H1~H' + str(len(_h_ranked)) + '=港股. ①N/②N=淘汰顺序. 灰色=已淘汰.', len(cols4_header))
    hdr_row(ws4, 3, cols4_header)
    EXEMPT_BLUE = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    ELIM_FILL = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    for idx, s in enumerate(top1_sorted):
        row = 4 + idx
        _cat = _cat_order(s['cg'])
        if _cat == 0:
            rank_val = _a_rank.get(s['code'], '-')
        elif _cat == 1:
            rank_val = _h_rank.get(s['code'], '-')
        else:
            rank_val = '-'
        vals = [s['code'], s['name'], s['cg'], s['amount_yi'], round(s.get('log_score',0),3),
                rank_val, _removed_order.get(s['code'], '')]
        elim_label = vals[-1]  # elimination order, empty if kept
        hnames = get_holding_names(s['code'])
        is_hk = s['cg'].startswith('HK-')
        if is_hk:
            hnames = [h + ' -H' for h in hnames]  # suffix HK holdings
        if _cat_order(s['cg']) == 2:
            hnames = []  # QDII/commodity: no comparable holdings
        vals += hnames + ['']*(10-len(hnames))
        for ci, val in enumerate(vals, 1):
            if isinstance(val, float): val = round(val, 2)
            c = ws4.cell(row=row, column=ci, value=val)
            c.font = Font(name="Consolas", size=9); c.border = thin_border
        if elim_label:
            for ci in range(1, len(cols4_header)+1):
                ws4.cell(row=row, column=ci).fill = ELIM_FILL
        elif _cat_order(s['cg']) == 2:
            for ci in range(1, len(cols4_header)+1):
                ws4.cell(row=row, column=ci).fill = EXEMPT_BLUE
        elif not elim_label:
            # Color duplicate holdings (only for kept ETFs)
            for ci in range(6, len(cols4_header)+1):
                val = ws4.cell(row=row, column=ci).value
                if val in _dup_fills:
                    ws4.cell(row=row, column=ci).fill = _dup_fills[val]

    # Remove bottom PK pairs table — now just a summary line
    rule_row(ws4, 6 + len(top1_sorted), f'淘汰 {_total_removed} 支: ①底1/2(O1>5)={len(removed_pass1)}支 + ②贪心(O2>=8)={len(removed_pass2)}支. 灰色=已淘汰.', len(cols4_header))

    # PK pairs summary now inline in main table (see 排除 column)

    widths4 = [10,30,22,14,10,8,8] + [10]*10
    for i, w in enumerate(widths4): ws4.column_dimensions[get_column_letter(i+1)].width = w

    # Sheet 5: Full diff — selected + missing pool, with status labels
    ws5 = wb.create_sheet("5-差异对照")
    NEW_FILL    = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')  # green: new
    POOL_FILL   = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # yellow: pool
    MISS_FILL   = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')  # red: missing
    REPL_FILL   = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')  # blue: replace
    NEW_FONT    = Font(name='Consolas', size=9, bold=True, color='065F46')

    cols5 = ['code','name','type','amount_yi','cg','status']
    _new_codes = sel_codes - pool_codes
    n_new = len(_new_codes)

    # Build elimination info for recommendation logic
    _elim_info = {}
    for _e in removed_pass1:
        _elim_info[_e] = ('①', _all_rank.get(_e, 0))
    for _e in removed_pass2:
        _elim_info[_e] = ('②', _all_rank.get(_e, 0))

    # Build score lookup for ALL ETFs (not just winners), for marginal-difference detection.
    _all_scores = {}
    for _, _r in df_filtered.iterrows():
        _all_scores[_r['code']] = _r.get('log_score', 0)
    _group_winner = {}
    for _s in top1_per_cg:
        _cg = _s['cg']
        if _cg not in _group_winner or _s.get('log_score', 0) > _group_winner[_cg][1]:
            _group_winner[_cg] = (_s['code'], _s.get('log_score', 0))

    # Score margin threshold: if challenger beats pool ETF by < MARGIN, don't recommend
    SCORE_MARGIN = 0.05

    # Build pool coarse-group set and holdings overlap lookup for rating
    _pool_cgs = set()
    for _pc in pool_codes:
        _prow = df[df['code'] == _pc]
        if len(_prow) > 0:
            _pool_cgs.add(classify(_prow.iloc[0]))

    # Max Jaccard overlap with any pool ETF, excluding self
    def _max_pool_jaccard(code):
        _sh = hs(code)
        if not _sh: return 0
        _max_j = 0
        for _pc in pool_codes:
            if _pc == code: continue  # exclude self
            _ph = hs(_pc)
            if not _ph: continue
            _j = len(_sh & _ph) / max(1, len(_sh | _ph))
            if _j > _max_j: _max_j = _j
        return _max_j

    # Data lookup helpers
    def _get(code, field, default=0):
        _rows = df_filtered[df_filtered['code'] == code]
        return _rows[field].values[0] if len(_rows) > 0 else default

    def _rating(score, reasons):
        """Map score 0-8 to 5-tier rating."""
        if score >= 6: return '强烈推荐', '; '.join(reasons)
        if score >= 4: return '推荐', '; '.join(reasons)
        if score >= 2: return '中立', '; '.join(reasons)
        if score >= 1: return '不推荐', '; '.join(reasons)
        return '强烈不推荐', '; '.join(reasons)

    def _recommend(status, code, amount_yi):
        """4-dimension rating: turnover, data, AUM, overlap."""
        amt = amount_yi if amount_yi else 0
        rows = _get(code, 'history_days', 0)
        aum = _get(code, 'mcap_yi', 0)
        cg = classify(df[df['code'] == code].iloc[0]) if len(df[df['code'] == code]) > 0 else ''
        overlap = _max_pool_jaccard(code)

        score = 4  # baseline
        reasons = []

        # Turnover (0–2)
        if amt >= 10: score += 2; reasons.append(f'高成交额{amt:.0f}亿')
        elif amt >= 5: score += 1; reasons.append(f'成交额{amt:.0f}亿')
        elif amt >= 1: pass
        elif amt >= 0.5: score -= 1; reasons.append(f'成交额偏低{amt:.2f}亿')
        else: score -= 2; reasons.append(f'成交额不足{amt:.2f}亿')

        # Data length (0–2)
        if rows >= 1000: score += 2; reasons.append(f'数据充足{rows}d')
        elif rows >= 500: score += 1
        elif rows >= 80: pass
        else: score -= 2; reasons.append(f'数据不足{rows}d')

        # Fund size (0–1)
        if aum >= 100: score += 1; reasons.append(f'大规模{aum:.0f}亿')
        elif aum >= 10: pass
        elif aum >= 1: score -= 1; reasons.append(f'规模偏小{aum:.1f}亿')
        else: score -= 2; reasons.append('规模极小')

        # Holdings overlap (0–2 penalty)
        if overlap > 0.6: score -= 2; reasons.append(f'成分股高度重叠{overlap:.0%}')
        elif overlap > 0.4: score -= 1; reasons.append(f'成分股部分重叠{overlap:.0%}')

        # Status-specific adjustments
        if status == '缺失':
            if amt < 0.3: return '强烈推荐', f'成交额{amt:.2f}亿<30M，应剔除'
            info = _elim_info.get(code)
            if info:
                label, rank = info
                if label == '①': return '推荐', f'底半区淘汰 rank={rank}'
                return '中立', f'贪心削峰淘汰 rank={rank}'
            _my_score = _all_scores.get(code, 0)
            if cg in _group_winner:
                _diff = _group_winner[cg][1] - _my_score
                if _diff < SCORE_MARGIN:
                    return '中立', f'差距微弱 Δ={_diff:.3f}'
                return _rating(score - 2, reasons + [f'PK落败 Δ={_diff:.3f}'])
            return _rating(score - 2, reasons)
        elif status == '新增':
            _replaces = None
            for _mc in missing:
                _mrow = df[df['code'] == _mc]
                if len(_mrow) > 0 and classify(_mrow.iloc[0]) == cg:
                    _replaces = _mc
                    break
            if _replaces:
                _diff = _all_scores.get(code, 0) - _all_scores.get(_replaces, 0)
                if _diff < SCORE_MARGIN:
                    return '中立', f'差距微弱 Δ={_diff:.3f}'
                reasons.append(f'优于池内{_replaces} Δ={_diff:.3f}')
            return _rating(score, reasons)
        return '—', '已覆盖'

    # Keep backward compat for combined list building
    def _recommend_simple(status, code, amount_yi):
        return _recommend(status, code, amount_yi)

    # Build combined list with replacement detection.
    # Match missing pool ETFs to new selections in the same coarse group → "替换"
    _cg_of_missing = {}  # code -> cg for missing pool ETFs
    for code in sorted(missing):
        _mrow = df[df['code'] == code]
        if len(_mrow) > 0:
            _cg_of_missing[code] = classify(_mrow.iloc[0])

    _replacement_pairs = []  # [(old_code, new_code, cg, old_score, new_score)]
    _replaced_news = set()
    _replaced_missings = set()
    for _mc, _mcg in _cg_of_missing.items():
        for _s in kept:
            if _s['code'] in _new_codes and _s['cg'] == _mcg:
                _replacement_pairs.append((_mc, _s['code'], _mcg,
                    _all_scores.get(_mc, 0), _s.get('log_score', 0)))
                _replaced_news.add(_s['code'])
                _replaced_missings.add(_mc)
                break  # one replacement per missing ETF

    _combined = []
    # Pool ETFs that stayed selected
    for s in kept:
        if s['code'] not in _new_codes:
            entry = dict(s)
            entry['status'] = '池内'
            entry['recommend'] = '—'
            entry['rec_reason'] = '已覆盖，无需操作'
            _combined.append(entry)
    # Replacements (merged rows)
    for _old, _new, _cg, _old_score, _new_score in _replacement_pairs:
        _new_s = [s for s in kept if s['code'] == _new][0]
        _old_row = df[df['code'] == _old].iloc[0] if len(df[df['code'] == _old]) > 0 else None
        rec, rec_reason = _recommend('新增', _new, _new_s.get('amount_yi', 0))
        entry = {
            'code': f'{_old}→{_new}',
            'name': f'{_old_row["name"] if _old_row is not None else _old} → {_new_s["name"]}',
            'type': _new_s.get('type', ''),
            'amount_yi': f'{_new_s["amount_yi"]} (旧 {_old_row["amount_yi"] if _old_row is not None else "?"})',
            'cg': _cg,
            'status': '替换',
            'recommend': rec,
            'rec_reason': rec_reason,
        }
        _combined.append(entry)
    # New ETFs that are pure additions (not replacing any pool ETF)
    for s in kept:
        if s['code'] in _new_codes and s['code'] not in _replaced_news:
            entry = dict(s)
            entry['status'] = '新增'
            rec, rec_reason = _recommend('新增', s['code'], s.get('amount_yi', 0))
            entry['recommend'] = rec
            entry['rec_reason'] = rec_reason
            _combined.append(entry)
    # Missing pool ETFs that weren't matched to a replacement
    for code in sorted(missing):
        if code not in _replaced_missings:
            _mrow = df[df['code'] == code]
            if len(_mrow) > 0:
                r = _mrow.iloc[0]
                rec, rec_reason = _recommend('缺失', code, r['amount_yi'])
                entry = {'code': code, 'name': r['name'], 'type': r['type'],
                         'amount_yi': r['amount_yi'], 'cg': classify(r), 'status': '缺失',
                         'recommend': rec, 'rec_reason': rec_reason}
                _combined.append(entry)

    _combined.sort(key=lambda s: ({'池内':0,'替换':1,'新增':2,'缺失':3}[s['status']], s['cg'], -float(str(s.get('amount_yi','0')).split()[0]) if s.get('amount_yi') else 0))

    # Rating color map: 5 tiers
    REC_FILLS = {
        '强烈推荐': PatternFill(start_color='065F46', end_color='065F46', fill_type='solid'),
        '推荐':     PatternFill(start_color='10B981', end_color='10B981', fill_type='solid'),
        '中立':     PatternFill(start_color='FCD34D', end_color='FCD34D', fill_type='solid'),
        '不推荐':   PatternFill(start_color='F97316', end_color='F97316', fill_type='solid'),
        '强烈不推荐': PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid'),
        '—':       None,
    }
    REC_FONTS = {
        '强烈推荐': Font(name='Consolas', size=9, bold=True, color='FFFFFF'),
        '推荐':     Font(name='Consolas', size=9, bold=True, color='FFFFFF'),
        '不推荐':   Font(name='Consolas', size=9, bold=True, color='FFFFFF'),
        '强烈不推荐': Font(name='Consolas', size=9, bold=True, color='FFFFFF'),
        '中立':     Font(name='Consolas', size=9, color='92400E'),
    }

    cols5 = ['code','name','type','amount_yi','cg','status','recommend']
    ncols5 = len(cols5)
    n_rec = {'强烈推荐':0,'推荐':0,'中立':0,'不推荐':0,'强烈不推荐':0}
    for s in _combined:
        r = s.get('recommend', '—')
        if r in n_rec: n_rec[r] += 1
    n_repl = sum(1 for s in _combined if s['status'] == '替换')
    desc5 = f'差异对照: {len(kept)}入选 | 替换{n_repl}(蓝) | 新增{n_new - len(_replaced_news)}(绿) | 已覆盖{len(overlap)}(黄) | 未入选{len(missing) - len(_replaced_missings)}(红)'
    rule_row(ws5, 1, desc5, 10)
    rule_row(ws5, 2, f'评级(4维,0-8分): 成交额≥10亿+2/≥5亿+1/<1亿-1/<0.5亿-2 | 数据≥1000d+2/≥500d+1/<80d-2 | 规模≥100亿+1/<10亿-1/<1亿-2 | 重叠度>60%-2/>40%-1. 总分≥6=强烈推荐 ≥4=推荐 ≥2=中立 ≥1=不推荐.', 10)
    ncols5 = 10
    cols5 = ['code','name','amount_yi','history_days','mcap_yi','overlap','score','cg','status','recommend']
    hdr_row(ws5, 3, ['代码','名称','成交额(亿)','数据量(d)','规模(亿)','重叠度','评分','粗组','状态','评级'])

    # Build dimension data for each entry
    for _e in _combined:
        _code = _e.get('code', '')
        _actual_code = _code.split('→')[-1].strip() if '→' in str(_code) else _code
        _e['history_days'] = _get(_actual_code, 'history_days', 0)
        _e['mcap_yi'] = round(_get(_actual_code, 'mcap_yi', 0), 1)
        _e['overlap'] = round(_max_pool_jaccard(_actual_code), 2)
        # Compute score using same 4-dimension formula as _recommend
        _amt = float(str(_e.get('amount_yi', '0')).split()[0]) if _e.get('amount_yi') else 0
        _rows = _e.get('history_days', 0)
        _aum = _e.get('mcap_yi', 0)
        _score = 4  # baseline
        if _amt >= 10: _score += 2
        elif _amt >= 5: _score += 1
        elif _amt >= 1: pass
        elif _amt >= 0.5: _score -= 1
        else: _score -= 2
        if _rows >= 1000: _score += 2
        elif _rows >= 500: _score += 1
        elif _rows >= 80: pass
        else: _score -= 2
        if _aum >= 100: _score += 1
        elif _aum >= 10: pass
        elif _aum >= 1: _score -= 1
        else: _score -= 2
        _ol = _e.get('overlap', 0)
        if _ol > 0.6: _score -= 2
        elif _ol > 0.4: _score -= 1
        _e['score'] = _score

    for idx, s in enumerate(_combined):
        row = 4 + idx
        vals = [s.get('code',''), s.get('name',''), s.get('amount_yi',''),
                s.get('history_days',''), s.get('mcap_yi',''), s.get('overlap',''),
                s.get('score',''), s.get('cg',''), s.get('status',''), s.get('recommend','')]
        for ci, val in enumerate(vals, 1):
            if isinstance(val, float): val = round(val, 2)
            c = ws5.cell(row=row, column=ci, value=val)
            c.font = Font(name="Consolas", size=9); c.border = thin_border
        if s['status'] == '新增':
            for ci in range(1, ncols5+1):
                ws5.cell(row=row, column=ci).fill = NEW_FILL
            ws5.cell(row=row, column=1).font = NEW_FONT
        elif s['status'] == '替换':
            for ci in range(1, ncols5+1):
                ws5.cell(row=row, column=ci).fill = REPL_FILL
            ws5.cell(row=row, column=1).font = NEW_FONT
        elif s['status'] == '池内':
            for ci in range(1, ncols5+1):
                ws5.cell(row=row, column=ci).fill = POOL_FILL
        elif s['status'] == '缺失':
            for ci in range(1, ncols5+1):
                ws5.cell(row=row, column=ci).fill = MISS_FILL
        # Color rating column (col 10)
        rec = s.get('recommend', '')
        rec_fill = REC_FILLS.get(rec)
        rec_font = REC_FONTS.get(rec)
        if rec_fill:
            ws5.cell(row=row, column=10).fill = rec_fill
        if rec_font:
            ws5.cell(row=row, column=10).font = rec_font

    widths5 = [16,40,14,12,12,8,6,22,10,12]
    for i, w in enumerate(widths5):
        ws5.column_dimensions[get_column_letter(i+1)].width = w

    os.system('taskkill /f /im excel.exe 2>nul')
    import time; time.sleep(1)
    wb.save(str(OUTPUT))
    print(f"  Saved: {OUTPUT}")
    os.system(f'start excel "{OUTPUT}"')

